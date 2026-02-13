#!/usr/bin/env python3
"""Generate the Euangelion Master Image Library Excel document."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Theme colors ──
TEHOM_BLACK = "1A1612"
GOD_IS_GOLD = "C19A6B"
SCROLL_WHITE = "F7F3ED"
DARK_BG = "2A2520"
HEADER_FONT_COLOR = "FFFFFF"

# ── Style helpers ──
header_font = Font(name="Inter", bold=True, size=11, color=HEADER_FONT_COLOR)
header_fill = PatternFill(start_color=TEHOM_BLACK, end_color=TEHOM_BLACK, fill_type="solid")
gold_font = Font(name="Inter", bold=True, size=11, color=TEHOM_BLACK)
gold_fill = PatternFill(start_color=GOD_IS_GOLD, end_color=GOD_IS_GOLD, fill_type="solid")
body_font = Font(name="Inter", size=10, color="333333")
body_font_light = Font(name="Inter", size=10, color="666666")
title_font = Font(name="Inter", bold=True, size=14, color=TEHOM_BLACK)
subtitle_font = Font(name="Inter", bold=True, size=12, color=GOD_IS_GOLD)
thin_border = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)
wrap_alignment = Alignment(wrap_text=True, vertical="top")
center_alignment = Alignment(horizontal="center", vertical="center")
center_wrap = Alignment(horizontal="center", vertical="top", wrap_text=True)


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border


def style_category_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = gold_font
        cell.fill = gold_fill
        cell.border = thin_border


def style_body_cell(ws, row, col, light=False):
    cell = ws.cell(row=row, column=col)
    cell.font = body_font_light if light else body_font
    cell.alignment = wrap_alignment
    cell.border = thin_border
    return cell


def auto_width(ws, min_width=10, max_width=50):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


# ══════════════════════════════════════════════════════════════
# CREATE WORKBOOK
# ══════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()


# ══════════════════════════════════════════════════════════════
# SHEET 1: EVERGREEN ILLUSTRATIONS
# ══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Evergreen Illustrations"
ws1.sheet_properties.tabColor = GOD_IS_GOLD

headers = ["#", "Category", "Image Name", "Description / Scene", "Use Cases / Tags", "Primary Size", "Aspect Ratio", "Formats", "Priority", "Status"]
col_widths = [5, 20, 30, 45, 45, 16, 12, 16, 10, 12]

for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

ws1.append(headers)
style_header_row(ws1, 1, len(headers))
ws1.freeze_panes = "A2"

illustrations = [
    # ── CREATION & NATURE ──
    ("Creation & Nature", [
        ("Light breaking through darkness / void", "Creation, new beginnings, hope, revelation, Genesis 1", "1920x1080", "16:9"),
        ("Starfield / night sky", "God's promises, Abrahamic covenant, wonder, vastness", "1920x1080", "16:9"),
        ("Sunrise over water", "New mercies, resurrection, renewal, Lamentations 3", "1920x1080", "16:9"),
        ("Sunset / golden hour landscape", "Rest, Sabbath, day ending, reflection, evening prayer", "1920x1080", "16:9"),
        ("Single tree in an open field", "Psalm 1, rootedness, solitude, spiritual growth", "1200x675", "16:9"),
        ("Ancient olive tree (gnarled, textured)", "Gethsemane, endurance, Israel, anointing", "1200x675", "16:9"),
        ("Vineyard / grapevines", "Abiding, fruitfulness, pruning, John 15", "1200x675", "16:9"),
        ("Wheat field at harvest", "Harvest, provision, parables, sowing and reaping", "1920x1080", "16:9"),
        ("Seeds in soil (cross-section close-up)", "Faith, mustard seed, parable of sower, beginnings", "600x600", "1:1"),
        ("Fig tree with/without fruit", "Judgment, seasons, spiritual barrenness, Mark 11", "1200x675", "16:9"),
        ("Wildflowers in a field", "Lilies of the field, God's provision, beauty, Matthew 6", "1200x675", "16:9"),
        ("Mountain peak above clouds", "Transfiguration, encounter with God, ascent, Sinai", "1920x1080", "16:9"),
        ("Desert / barren wilderness", "Testing, wandering, 40 days, spiritual dryness", "1920x1080", "16:9"),
        ("Oasis in desert", "Refreshment, Psalm 23, living water, restoration", "1200x675", "16:9"),
        ("Stormy sea / crashing waves", "Trials, chaos, Jonah, Jesus calming storm", "1920x1080", "16:9"),
        ("Still waters (glassy, reflective)", "Psalm 23, peace, restoration, quiet trust", "1920x1080", "16:9"),
        ("River flowing through landscape", "Living water, baptism, Ezekiel's river, cleansing", "1920x1080", "16:9"),
        ("Rain falling on dry earth", "Blessing, Holy Spirit outpouring, renewal, refreshing", "1200x675", "16:9"),
        ("Rock / massive stone formation", "God as rock, foundation, Peter, cornerstone", "1200x675", "16:9"),
        ("Burning bush (fire that doesn't consume)", "God's call, holy ground, Moses, Exodus 3", "1200x675", "16:9"),
        ("Pillar of cloud / pillar of fire", "God's guidance, presence, wilderness journey", "1200x675", "16:9"),
        ("Rainbow after storm", "Covenant, promise, faithfulness, Noah, Genesis 9", "1920x1080", "16:9"),
        ("Single dove in flight", "Holy Spirit, peace, Noah, baptism of Jesus", "800x600", "4:3"),
        ("Lamb (young, white)", "Sacrifice, Jesus as Lamb, Passover, innocence", "800x600", "4:3"),
        ("Lion (majestic, still)", "Lion of Judah, courage, authority, Revelation 5", "800x600", "4:3"),
        ("Eagle soaring", "Isaiah 40, renewal of strength, vision, freedom", "1200x675", "16:9"),
    ]),

    # ── LIGHT & DARKNESS ──
    ("Light & Darkness", [
        ("Single candle in darkness", "Hope, witness, Word as lamp, Psalm 119", "800x600", "4:3"),
        ("Oil lamp (ancient clay)", "Parable of virgins, readiness, Psalm 119:105", "600x400", "3:2"),
        ("Shaft of light through crack/window", "Revelation, divine interruption, grace breaking in", "1200x675", "16:9"),
        ("Light piercing storm clouds", "God breaking through, deliverance, hope in trials", "1920x1080", "16:9"),
        ("Shadow and light on a face (chiaroscuro)", "Inner struggle, conviction, transformation, identity", "800x600", "4:3"),
        ("Torch / fire in darkness", "Guidance, Pentecost, truth, John 1 light", "800x600", "4:3"),
        ("Dawn breaking over a valley", "Deliverance, weeping to joy, Psalm 30:5", "1920x1080", "16:9"),
        ("Moonlight on still water", "Night season, quiet trust, vigil, waiting on God", "1920x1080", "16:9"),
        ("Lighthouse on rocky coast", "Guidance, warning, Christ as light, safety", "1200x675", "16:9"),
        ("Bonfire / campfire (warm glow)", "Community, fellowship, Peter's denial, warmth", "800x600", "4:3"),
        ("Stars against deep black sky", "Promises, Abraham, God's faithfulness, infinite", "1920x1080", "16:9"),
        ("Eclipse / sun being covered", "Crucifixion darkness, judgment, mystery, awe", "1200x675", "16:9"),
    ]),

    # ── SACRED OBJECTS & SYMBOLS ──
    ("Sacred Objects & Symbols", [
        ("Open ancient scroll", "Scripture, Word of God, Torah, revelation, study", "600x400", "3:2"),
        ("Closed scroll with seal", "Revelation, mystery, apocalyptic, sealed prophecy", "600x400", "3:2"),
        ("Open Bible / leather-bound book", "Study, devotion, quiet time, Word of God", "800x600", "4:3"),
        ("Hebrew text close-up (calligraphy)", "Word studies, original languages, Torah, beauty", "600x400", "3:2"),
        ("Greek manuscript text", "New Testament, word studies, manuscripts, history", "600x400", "3:2"),
        ("Bread (artisan loaf, broken)", "Communion, Last Supper, daily bread, manna, provision", "600x600", "1:1"),
        ("Wine / cup (chalice, goblet)", "Communion, blood of Christ, new covenant", "600x600", "1:1"),
        ("Bread and wine together", "Eucharist, Lord's Supper, remembrance, covenant meal", "800x600", "4:3"),
        ("Wooden cross (rugged, weathered)", "Crucifixion, sacrifice, gospel core, atonement", "1200x675", "16:9"),
        ("Empty cross silhouette against sky", "Resurrection, victory, hope, Easter", "1920x1080", "16:9"),
        ("Crown of thorns", "Suffering, kingship inverted, Passion, mockery", "600x600", "1:1"),
        ("Crown of thorns with gold light", "Suffering redeemed, kingship, glory through pain", "600x600", "1:1"),
        ("Nails (iron, rough)", "Crucifixion, cost of grace, sacrifice, Colossians 2", "600x400", "3:2"),
        ("Empty tomb / stone rolled away", "Resurrection, Easter, victory over death", "1920x1080", "16:9"),
        ("Linen burial cloth (folded)", "Resurrection detail, John 20, intentionality of God", "600x400", "3:2"),
        ("Shepherd's staff / rod", "Psalm 23, guidance, protection, pastoral leadership", "400x500", "4:5"),
        ("Clay jar / earthen vessel", "2 Cor 4, fragility, treasure within, humility", "600x600", "1:1"),
        ("Potter's hands shaping clay", "God as potter, formation, Isaiah 64, Jeremiah 18", "800x600", "4:3"),
        ("Anointing oil being poured", "Holy Spirit, consecration, healing, Psalm 23", "600x600", "1:1"),
        ("Incense rising (smoke curling upward)", "Prayer ascending, worship, temple, Revelation 8", "800x600", "4:3"),
        ("Menorah (seven branches, lit)", "God's presence, temple, light of world, Exodus 25", "600x400", "3:2"),
        ("Shofar (ram's horn)", "Call to worship, Jubilee, warning, return of Christ", "600x400", "3:2"),
        ("Altar of stones", "Sacrifice, worship, memorial, remembrance, Ebenezer", "800x600", "4:3"),
        ("Baptismal water (immersion)", "Baptism, death and rebirth, cleansing, Romans 6", "1200x675", "16:9"),
        ("Basin and towel", "Foot washing, servant leadership, humility, John 13", "600x400", "3:2"),
        ("Sackcloth and ashes", "Repentance, mourning, lament, Joel 2", "600x400", "3:2"),
        ("Gold refined by fire", "Purification, testing, 1 Peter 1:7, Malachi 3", "600x600", "1:1"),
        ("Armor pieces (helmet, shield, sword)", "Ephesians 6, spiritual warfare, readiness", "800x600", "4:3"),
        ("Sword (double-edged)", "Word of God, Hebrews 4:12, truth, discernment", "400x500", "4:5"),
        ("Key (ancient, ornate)", "Authority, Kingdom keys, Peter, Matthew 16", "400x500", "4:5"),
        ("Chain links (broken)", "Freedom, deliverance, liberation, Galatians 5", "600x400", "3:2"),
        ("Yoke (wooden, for oxen)", "Rest, Jesus' easy yoke, discipleship, Matthew 11", "600x400", "3:2"),
        ("Scales / balance", "Justice, judgment, righteousness, Proverbs 16", "600x600", "1:1"),
        ("Cornerstone / foundation stone", "Christ as cornerstone, building on rock, Ephesians 2", "600x400", "3:2"),
        ("Pearl (single, luminous)", "Pearl of great price, Kingdom value, Matthew 13", "600x600", "1:1"),
        ("Treasure chest / hidden treasure", "Kingdom of heaven parables, discovery, joy", "600x400", "3:2"),
        ("Salt crystals (close-up)", "Salt of the earth, preservation, flavor, Matthew 5", "600x600", "1:1"),
    ]),

    # ── HANDS & HUMAN FIGURES ──
    ("Hands & Human Figures", [
        ("Hands open, palms up (receiving)", "Surrender, prayer, openness, receiving grace", "800x600", "4:3"),
        ("Hands clasped in prayer", "Prayer, intercession, devotion, supplication", "800x600", "4:3"),
        ("Hands raised in worship", "Praise, surrender, adoration, Psalm 63", "1200x675", "16:9"),
        ("Weathered hands holding bread", "Communion, provision, elderly faith, endurance", "600x600", "1:1"),
        ("Hand reaching upward into light", "Crying out, desperation, seeking God, Psalm 42", "800x600", "4:3"),
        ("Two hands reaching toward each other", "Reconciliation, community, God reaching man", "1200x675", "16:9"),
        ("Hand writing with pen/quill", "Journaling, Scripture writing, reflection, study", "600x400", "3:2"),
        ("Hands holding soil with seedling", "Nurture, growth, stewardship, discipleship", "600x600", "1:1"),
        ("Hands washing another's feet", "Servanthood, humility, John 13, leadership", "800x600", "4:3"),
        ("Single figure walking a path", "Journey, pilgrimage, following God, faith walk", "1920x1080", "16:9"),
        ("Figure kneeling in prayer", "Submission, prayer, devotion, Gethsemane", "800x600", "4:3"),
        ("Figure standing at crossroads", "Decision, discernment, choosing God's way", "1200x675", "16:9"),
        ("Figure looking out at vast landscape", "Calling, vision, future, trust, Abraham", "1920x1080", "16:9"),
        ("Figure in doorway (light beyond)", "Threshold, invitation, 'I am the door,' John 10", "800x600", "4:3"),
        ("Shepherd with flock (silhouette)", "Psalm 23, leadership, care, John 10, pastoral", "1920x1080", "16:9"),
        ("Two people walking together", "Discipleship, Emmaus road, companionship, mentorship", "1200x675", "16:9"),
        ("Small gathering / circle of people", "Community, church, fellowship, Acts 2, koinonia", "1200x675", "16:9"),
        ("Person sitting alone in silence", "Solitude, contemplation, listening, stillness", "800x600", "4:3"),
        ("Person weeping / head bowed", "Lament, grief, brokenness, 'Jesus wept,' John 11", "800x600", "4:3"),
        ("Person with arms outstretched (cruciform)", "Surrender, Christlikeness, sacrifice, worship", "1200x675", "16:9"),
    ]),

    # ── ARCHITECTURE & SPACES ──
    ("Architecture & Spaces", [
        ("Ancient stone doorway / gate", "Narrow gate, entering God's presence, Matthew 7", "800x600", "4:3"),
        ("Narrow path through rocky terrain", "Narrow way, discipleship cost, Matthew 7:14", "1200x675", "16:9"),
        ("Wide road vs. narrow road (fork)", "Two paths, choices, wisdom, Proverbs, Matthew 7", "1200x675", "16:9"),
        ("Temple interior (columns, light)", "Worship, God's dwelling, holiness, awe", "1920x1080", "16:9"),
        ("Ruins of ancient temple/city", "Judgment, exile, rebuilding, Nehemiah, lament", "1200x675", "16:9"),
        ("Stone wall being rebuilt", "Restoration, Nehemiah, repairing breaches, renewal", "1200x675", "16:9"),
        ("Upper room (simple, table, warm light)", "Last Supper, Pentecost, fellowship, intimacy", "1200x675", "16:9"),
        ("Garden (lush, enclosed)", "Eden, Gethsemane, Song of Solomon, paradise", "1200x675", "16:9"),
        ("Walled city on a hill", "Jerusalem, city on a hill, Zion, witness", "1920x1080", "16:9"),
        ("Empty throne room", "God's sovereignty, Revelation 4, authority, worship", "1920x1080", "16:9"),
        ("Open door with light streaming through", "Invitation, opportunity, Rev 3:20, welcome", "800x600", "4:3"),
        ("Closed/locked door", "Persistence in prayer, waiting, patience, Luke 11", "800x600", "4:3"),
        ("Bridge over chasm", "Reconciliation, Christ as mediator, 1 Timothy 2", "1200x675", "16:9"),
        ("Well / cistern", "Living water, Samaritan woman, Jacob's well, John 4", "600x600", "1:1"),
        ("Watchtower", "Vigilance, prophetic watching, Habakkuk 2:1", "800x600", "4:3"),
        ("Vineyard wall / winepress", "Kingdom parables, judgment, harvest, Isaiah 5", "1200x675", "16:9"),
        ("Ancient library / scrolls on shelves", "Wisdom, study, Proverbs, knowledge, learning", "1200x675", "16:9"),
    ]),

    # ── TEXTURES & ABSTRACT ──
    ("Textures & Abstract", [
        ("Cracked, dry earth", "Spiritual thirst, drought, need for God, Psalm 63", "1200x675", "16:9"),
        ("Water ripples from single drop", "Impact, small beginnings, Word landing, faith", "600x600", "1:1"),
        ("Smoke / mist rising", "Mystery, God's presence, incense, prayer, temple", "800x600", "4:3"),
        ("Gold leaf texture on dark surface", "Glory, divinity, sacred, Tehom+Gold brand aesthetic", "512x512", "1:1"),
        ("Rough hewn wood grain", "Cross, carpenter, Jesus' trade, simplicity, craft", "512x512", "1:1"),
        ("Ancient parchment / papyrus texture", "Scripture, antiquity, written Word, Torah", "512x512", "1:1"),
        ("Shattered pottery pieces", "Brokenness, being remade, Jeremiah 18, restoration", "600x600", "1:1"),
        ("Woven fabric / threads", "Unity, body of Christ, interconnection, 1 Cor 12", "512x512", "1:1"),
        ("Thorns close-up", "Fall, curse, suffering, crown of thorns, Genesis 3", "600x400", "3:2"),
        ("Ember / glowing coal", "Purification, Isaiah 6, holy fire, zeal", "600x600", "1:1"),
        ("Ash / dust", "Mortality, repentance, dust to dust, Genesis 3:19", "600x600", "1:1"),
        ("Oil on water (iridescent)", "Anointing, Holy Spirit, Psalm 133, unity", "600x600", "1:1"),
        ("Frost / ice melting", "Thawing of heart, renewal, spring, new season", "600x600", "1:1"),
        ("Sandal prints in dust", "Following, walking with God, pilgrimage, journey", "1200x675", "16:9"),
        ("Fingerprint close-up", "Identity, uniquely made, imago Dei, Psalm 139", "600x600", "1:1"),
        ("Mosaic tiles (ancient, fragmented)", "Beauty from pieces, restoration, God's artistry", "600x600", "1:1"),
    ]),

    # ── SEASONS & TIME ──
    ("Seasons & Time", [
        ("Spring blossoms on bare branch", "New life, resurrection, renewal, Song of Solomon", "1200x675", "16:9"),
        ("Summer abundance / full harvest", "Blessing, fruitfulness, provision, Galatians 6", "1200x675", "16:9"),
        ("Autumn leaves falling", "Letting go, seasons changing, release, Ecclesiastes", "1200x675", "16:9"),
        ("Winter bare tree (stark, beautiful)", "Dormancy, waiting, hidden growth, patience", "1200x675", "16:9"),
        ("Hourglass / sand falling", "Time, kairos vs chronos, urgency, Ecclesiastes 3", "600x600", "1:1"),
        ("Sundial casting shadow", "Appointed times, God's timing, patience", "600x600", "1:1"),
        ("Morning dew on grass", "Fresh mercy, Lamentations 3:23, new every morning", "1200x675", "16:9"),
        ("Twilight / between times", "Transition, liminality, waiting, threshold", "1920x1080", "16:9"),
    ]),

    # ── JOURNEY & MOVEMENT ──
    ("Journey & Movement", [
        ("Footprints on a path", "Following God, journey, discipleship, walk of faith", "1200x675", "16:9"),
        ("Road stretching into distance", "Calling, pilgrimage, future trust, Hebrews 11", "1920x1080", "16:9"),
        ("Mountain ascent trail", "Growth, difficulty, perseverance, pressing on", "1200x675", "16:9"),
        ("Valley between mountains", "Valley of shadow, humility, low seasons, Psalm 23", "1920x1080", "16:9"),
        ("Boat on open water", "Faith, risk, mission, Peter walking on water", "1200x675", "16:9"),
        ("Anchor (heavy, iron)", "Hope as anchor, Hebrews 6:19, steadfastness", "600x600", "1:1"),
        ("Open gate in a wall", "Freedom, invitation, access to God, Hebrews 10", "800x600", "4:3"),
        ("Footbridge over stream", "Crossing over, transition, faith step, Jordan", "1200x675", "16:9"),
        ("Compass on old map", "Guidance, direction, God's leading, Proverbs 3:5-6", "600x600", "1:1"),
        ("Ship in a storm", "Trials, Acts 27, faith under pressure, endurance", "1200x675", "16:9"),
        ("Sail catching wind", "Holy Spirit (ruach), being carried, mission, Acts", "1200x675", "16:9"),
    ]),

    # ── MAJOR BIBLICAL SCENES ──
    ("Major Biblical Scenes", [
        ("Creation — light separating from dark", "Genesis 1, beginnings, God's power, creation week", "1920x1080", "16:9"),
        ("Garden of Eden (lush, luminous)", "Innocence, original design, paradise, Genesis 2", "1920x1080", "16:9"),
        ("The Fall — hand reaching for fruit", "Sin, temptation, consequence, Genesis 3", "1200x675", "16:9"),
        ("Noah's ark on floodwaters", "Judgment + salvation, obedience, faith, Genesis 6-9", "1920x1080", "16:9"),
        ("Abraham under stars", "Promise, faith, covenant, Genesis 15, Hebrews 11", "1920x1080", "16:9"),
        ("Jacob wrestling the angel", "Struggling with God, transformation, Genesis 32", "1200x675", "16:9"),
        ("Moses at the Red Sea (parting)", "Deliverance, impossible odds, God's power, Exodus 14", "1920x1080", "16:9"),
        ("Moses on Sinai (tablets, fire)", "Law, covenant, encounter with God, Exodus 19-20", "1920x1080", "16:9"),
        ("Pillar of fire in the wilderness", "God's presence, guidance, protection, Exodus 13", "1200x675", "16:9"),
        ("David and Goliath (moment before)", "Faith vs fear, unlikely victory, 1 Samuel 17", "1200x675", "16:9"),
        ("David playing harp / lyre", "Worship, Psalms, intimacy with God, 1 Samuel 16", "800x600", "4:3"),
        ("Elijah on Mount Carmel (fire from heaven)", "God's power, confrontation, boldness, 1 Kings 18", "1920x1080", "16:9"),
        ("Elijah in the cave (still small voice)", "Hearing God, burnout, gentleness, 1 Kings 19", "800x600", "4:3"),
        ("Daniel in the lions' den", "Faithfulness under threat, deliverance, Daniel 6", "1200x675", "16:9"),
        ("Jonah and the great fish", "Running from God, mercy, second chances, Jonah 1-2", "1200x675", "16:9"),
        ("Isaiah's vision (throne room, seraphim)", "Holiness, calling, 'Here am I,' Isaiah 6", "1920x1080", "16:9"),
        ("Ezekiel's valley of dry bones", "Resurrection, restoration, impossible hope, Ezek 37", "1920x1080", "16:9"),
        ("Nativity — manger with light", "Incarnation, humility, Emmanuel, Luke 2", "1200x675", "16:9"),
        ("Magi following the star", "Seeking Jesus, journey of faith, Matthew 2", "1200x675", "16:9"),
        ("Jesus' baptism (water + dove + light)", "Identity, commissioning, Trinity, Matthew 3", "1200x675", "16:9"),
        ("Jesus in the wilderness (40 days)", "Temptation, endurance, spiritual battle, Matthew 4", "1200x675", "16:9"),
        ("Sermon on the Mount (crowd, hillside)", "Teaching, Beatitudes, kingdom ethics, Matthew 5-7", "1920x1080", "16:9"),
        ("Jesus calming the storm", "Authority, peace, 'be still,' Mark 4", "1200x675", "16:9"),
        ("Jesus walking on water", "Faith, impossibility, trust, Matthew 14", "1200x675", "16:9"),
        ("Feeding the 5,000 (loaves and fish)", "Abundance, provision, miracle, John 6", "1200x675", "16:9"),
        ("Jesus with children", "Welcome, innocence, kingdom belonging, Mark 10", "800x600", "4:3"),
        ("The Transfiguration (glowing figure)", "Glory revealed, divine nature, Matthew 17", "1920x1080", "16:9"),
        ("Jesus weeping over Jerusalem", "Compassion, lament, grief for lost, Luke 19", "800x600", "4:3"),
        ("Last Supper table", "Covenant, communion, betrayal, fellowship, Luke 22", "1920x1080", "16:9"),
        ("Gethsemane (agonized prayer)", "Surrender, suffering, 'not my will,' Luke 22", "1200x675", "16:9"),
        ("Crucifixion (three crosses, dark sky)", "Atonement, sacrifice, love, salvation, John 19", "1920x1080", "16:9"),
        ("Pierced hands on the cross", "Cost of grace, wounds, redemption, Isaiah 53", "800x600", "4:3"),
        ("Empty tomb at dawn", "Resurrection, victory, hope, Easter, Matthew 28", "1920x1080", "16:9"),
        ("Road to Emmaus (two figures + stranger)", "Encounter, revelation, burning hearts, Luke 24", "1200x675", "16:9"),
        ("Pentecost (tongues of flame)", "Holy Spirit, empowerment, church birth, Acts 2", "1920x1080", "16:9"),
        ("Paul on the Damascus Road (blinding light)", "Conversion, calling, dramatic change, Acts 9", "1200x675", "16:9"),
        ("New Jerusalem descending", "Consummation, hope, Revelation 21, new creation", "1920x1080", "16:9"),
    ]),

    # ── EMOTIONAL / EXPERIENTIAL ──
    ("Emotional / Experiential", [
        ("Tears on a face (close-up)", "Grief, lament, compassion, brokenness, sorrow", "600x600", "1:1"),
        ("Embrace / hug (two figures)", "Prodigal return, reconciliation, comfort, Luke 15", "800x600", "4:3"),
        ("Child sleeping peacefully", "Trust, rest, childlike faith, Psalm 131", "800x600", "4:3"),
        ("Person laughing with genuine joy", "Joy, celebration, abundant life, Nehemiah 8:10", "800x600", "4:3"),
        ("Eyes closed in contemplation", "Meditation, inner life, prayer, stillness", "600x600", "1:1"),
        ("Face looking upward into light", "Hope, seeking God, transformation, 2 Cor 3:18", "800x600", "4:3"),
        ("Scars on hands (healed but visible)", "Wounds redeemed, testimony, resurrection body", "600x600", "1:1"),
        ("Tears becoming streams of water", "Sorrow to joy, Psalm 126, redemptive suffering", "600x600", "1:1"),
    ]),

    # ── COMMUNITY & CHURCH ──
    ("Community & Church", [
        ("Breaking bread together (table scene)", "Fellowship, communion, hospitality, Acts 2:42", "1200x675", "16:9"),
        ("Diverse hands joined together", "Unity, body of Christ, reconciliation, Eph 4", "800x600", "4:3"),
        ("Baptism scene (river or pool)", "New life, public faith, belonging, Romans 6", "1200x675", "16:9"),
        ("Congregation singing / worship", "Corporate worship, praise, Psalm 150", "1920x1080", "16:9"),
        ("Mentor and younger person talking", "Discipleship, Titus 2, teaching, passing faith on", "800x600", "4:3"),
        ("Sharing a meal at a long table", "Feast, kingdom banquet, hospitality, Luke 14", "1920x1080", "16:9"),
        ("Candlelit gathering (intimate)", "House church, early church, Acts 2, koinonia", "1200x675", "16:9"),
        ("Hands anointing someone's head", "Ordination, blessing, healing prayer, James 5", "600x600", "1:1"),
    ]),
]

row_num = 2
img_num = 1
for category, items in illustrations:
    # Category row
    ws1.cell(row=row_num, column=1, value="")
    ws1.cell(row=row_num, column=2, value=category.upper())
    for c in range(1, len(headers) + 1):
        ws1.cell(row=row_num, column=c)
    style_category_row(ws1, row_num, len(headers))
    row_num += 1

    for name, use_cases, size, ratio in items:
        ws1.cell(row=row_num, column=1, value=img_num)
        ws1.cell(row=row_num, column=2, value=category)
        ws1.cell(row=row_num, column=3, value=name)
        ws1.cell(row=row_num, column=4, value=name)
        ws1.cell(row=row_num, column=5, value=use_cases)
        ws1.cell(row=row_num, column=6, value=size)
        ws1.cell(row=row_num, column=7, value=ratio)
        ws1.cell(row=row_num, column=8, value="WebP + JPEG")
        ws1.cell(row=row_num, column=9, value="")
        ws1.cell(row=row_num, column=10, value="To Do")
        for c in range(1, len(headers) + 1):
            style_body_cell(ws1, row_num, c)
        ws1.cell(row=row_num, column=1).alignment = center_alignment
        ws1.cell(row=row_num, column=7).alignment = center_wrap
        ws1.cell(row=row_num, column=10).alignment = center_wrap
        img_num += 1
        row_num += 1


# ══════════════════════════════════════════════════════════════
# SHEET 2: IMAGE SIZE SPECIFICATIONS
# ══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Image Size Specs")
ws2.sheet_properties.tabColor = "8B4513"

headers2 = ["Image Type", "Variant", "Dimensions (px)", "Aspect Ratio", "Format", "Retina (@2x)", "Max File Size", "Notes"]
col_widths2 = [22, 28, 18, 14, 14, 14, 14, 40]
for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.append(headers2)
style_header_row(ws2, 1, len(headers2))
ws2.freeze_panes = "A2"

size_specs = [
    # Heroes
    ("HERO IMAGES", None),
    ("Hero", "Desktop", "1920 x 1080", "16:9", "WebP + JPEG", "3840 x 2160", "250KB / 500KB", "Full-width landing, series headers"),
    ("Hero", "Tablet", "1024 x 768", "4:3", "WebP + JPEG", "2048 x 1536", "150KB / 300KB", "Tablet breakpoint"),
    ("Hero", "Mobile", "750 x 1000", "3:4", "WebP + JPEG", "1500 x 2000", "150KB / 300KB", "Mobile full-bleed"),
    ("Hero", "Mobile Tall (Story)", "750 x 1334", "9:16", "WebP + JPEG", "1500 x 2668", "200KB / 400KB", "Immersive story-style mobile"),

    # Cards
    ("CARD THUMBNAILS", None),
    ("Card", "Large", "800 x 600", "4:3", "WebP", "1600 x 1200", "80KB / 160KB", "Featured devotional cards"),
    ("Card", "Medium (Square)", "600 x 600", "1:1", "WebP", "1200 x 1200", "60KB / 120KB", "Grid cards, series overview"),
    ("Card", "Small", "400 x 300", "4:3", "WebP", "800 x 600", "40KB / 80KB", "Compact list cards"),
    ("Card", "Strip", "800 x 200", "4:1", "WebP", "1600 x 400", "50KB / 100KB", "Horizontal list items"),

    # Inline
    ("INLINE CONTENT", None),
    ("Inline", "Full-width", "1200 x 675", "16:9", "WebP + JPEG", "2400 x 1350", "150KB / 300KB", "Within devotional body text"),
    ("Inline", "Half-width", "600 x 600", "1:1", "WebP", "1200 x 1200", "60KB / 120KB", "Floated beside text"),
    ("Inline", "Word Study Card", "600 x 400", "3:2", "WebP", "1200 x 800", "50KB / 100KB", "Hebrew/Greek vocab modules"),
    ("Inline", "Scripture Block", "1200 x 480", "5:2", "WebP", "2400 x 960", "100KB / 200KB", "Anchor verse background"),
    ("Inline", "Profile Card", "400 x 500", "4:5", "WebP", "800 x 1000", "50KB / 100KB", "Biblical character portraits"),

    # Social
    ("SOCIAL / SEO / SHARE", None),
    ("Social", "OG Image (Facebook/LinkedIn)", "1200 x 630", "1.91:1", "JPEG", "N/A", "100KB", "Link previews, iMessage"),
    ("Social", "Twitter/X Card", "1200 x 628", "~1.91:1", "JPEG", "N/A", "100KB", "Large summary card"),
    ("Social", "Instagram Square", "1080 x 1080", "1:1", "JPEG", "N/A", "200KB", "Feed posts"),
    ("Social", "Instagram/TikTok Story", "1080 x 1920", "9:16", "JPEG", "N/A", "250KB", "Stories format"),
    ("Social", "Pinterest Pin", "1000 x 1500", "2:3", "JPEG", "N/A", "200KB", "Tall pin format"),
    ("Social", "WhatsApp Thumbnail", "300 x 300", "1:1", "JPEG", "N/A", "30KB", "Chat link preview"),

    # Textures
    ("BACKGROUND TEXTURES", None),
    ("Texture", "Parchment", "512 x 512", "1:1", "WebP", "N/A", "50KB", "Seamlessly tileable, Scripture areas"),
    ("Texture", "Dark Noise/Grain", "256 x 256", "1:1", "PNG", "N/A", "20KB", "Subtle body background overlay"),
    ("Texture", "Gold Leaf", "512 x 512", "1:1", "WebP", "N/A", "50KB", "Accent areas, premium feel"),
    ("Texture", "Linen", "256 x 256", "1:1", "PNG", "N/A", "20KB", "Card backgrounds"),
    ("Texture", "Stone", "512 x 512", "1:1", "WebP", "N/A", "50KB", "Headers, divider areas"),
    ("Texture", "Papyrus", "512 x 512", "1:1", "WebP", "N/A", "50KB", "Word study module backgrounds"),

    # Empty States
    ("EMPTY STATE / SYSTEM ILLUSTRATIONS", None),
    ("Empty State", "All states", "400 x 400", "1:1", "SVG (PNG @2x)", "800 x 800", "30KB SVG", "Line art: Scroll White + Gold on Tehom Black"),

    # Onboarding
    ("ONBOARDING ILLUSTRATIONS", None),
    ("Onboarding", "All slides", "600 x 600", "1:1", "SVG or WebP", "1200 x 1200", "50KB", "Branded illustration style"),

    # Placeholders
    ("LOADING / PLACEHOLDERS", None),
    ("Placeholder", "BlurHash / LQIP", "32 x 32", "Varies", "Base64 inline", "N/A", "<1KB", "Inline in HTML, per image"),
    ("Placeholder", "Skeleton Card", "400 x 300", "4:3", "SVG", "N/A", "5KB", "Animated shimmer skeleton"),
    ("Placeholder", "Generic Devotional", "1200 x 675", "16:9", "WebP", "N/A", "50KB", "Fallback before image loads"),
    ("Placeholder", "Generic Series", "800 x 600", "4:3", "WebP", "N/A", "40KB", "Missing series art fallback"),
]

row_num2 = 2
for item in size_specs:
    if item[1] is None:
        # Category row
        ws2.cell(row=row_num2, column=1, value=item[0])
        style_category_row(ws2, row_num2, len(headers2))
        row_num2 += 1
    else:
        for col_idx, val in enumerate(item, 1):
            ws2.cell(row=row_num2, column=col_idx, value=val)
            style_body_cell(ws2, row_num2, col_idx)
        row_num2 += 1


# ══════════════════════════════════════════════════════════════
# SHEET 3: EMPTY STATES & SYSTEM ILLUSTRATIONS
# ══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Empty States & System")
ws3.sheet_properties.tabColor = "4A4A4A"

headers3 = ["#", "Illustration", "Description / Scene", "Dimensions", "Format", "Where Used", "Status"]
col_widths3 = [5, 30, 45, 16, 16, 35, 12]
for i, w in enumerate(col_widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.append(headers3)
style_header_row(ws3, 1, len(headers3))
ws3.freeze_panes = "A2"

empty_states = [
    ("EMPTY STATE ILLUSTRATIONS", None),
    (1, "Empty scroll (unfurled, blank)", "No devotionals saved or found yet", "400x400", "SVG + PNG @2x", "Empty reading list, bookmarks"),
    (2, "Desert landscape (vast, quiet)", "Search returned no results", "400x400", "SVG + PNG @2x", "Empty search results page"),
    (3, "Cloud with X mark", "Device is offline", "400x400", "SVG + PNG @2x", "PWA offline fallback page"),
    (4, "Hourglass with sand flowing", "Content is loading, please wait", "400x400", "SVG + PNG @2x", "Skeleton/loading states"),
    (5, "Cracked clay vessel", "Something went wrong", "400x400", "SVG + PNG @2x", "500 error page"),
    (6, "Lost sheep looking around", "Page not found", "400x400", "SVG + PNG @2x", "404 error page"),
    (7, "Open door with light streaming in", "Welcome to your first visit", "400x400", "SVG + PNG @2x", "Onboarding first screen"),
    (8, "Crown / olive wreath", "You completed a series!", "400x400", "SVG + PNG @2x", "Series completion celebration"),
    (9, "Rising flame (steady, warm)", "Devotional streak is going!", "400x400", "SVG + PNG @2x", "Streak tracking UI"),
    (10, "Ribbon on scroll", "Devotional saved successfully", "200x200", "SVG + PNG @2x", "Bookmark/save confirmation toast"),
    (11, "Compass with golden needle", "Let's find your pathway", "400x400", "SVG + PNG @2x", "Soul Audit intro screen"),
    (12, "Broken chain link", "No internet connection", "400x400", "SVG + PNG @2x", "Connection error state"),

    ("ONBOARDING ILLUSTRATIONS", None),
    (13, "Scroll merging with phone screen", "Ancient wisdom, modern design", "600x600", "SVG or WebP @2x", "Onboarding slide 1"),
    (14, "Three nested circles (1/5/15)", "Three reading depths explained", "600x600", "SVG or WebP @2x", "Onboarding slide 2"),
    (15, "Three diverging roads (moon/sun/staff)", "Choose your pathway", "600x600", "SVG or WebP @2x", "Onboarding slide 3 — Sleep/Awake/Shepherd"),
    (16, "Sun cycle (dawn to dusk)", "Build a daily rhythm", "600x600", "SVG or WebP @2x", "Onboarding slide 4"),
    (17, "Reflection in still water (mirror)", "Soul Audit assessment", "600x600", "SVG or WebP @2x", "Soul Audit intro"),

    ("SUCCESS / ACHIEVEMENT ILLUSTRATIONS", None),
    (18, "Sunrise over mountain summit", "Series milestone reached", "400x400", "SVG + PNG @2x", "Mid-series milestone"),
    (19, "Tree bearing fruit", "Consistent daily reading", "400x400", "SVG + PNG @2x", "7-day streak celebration"),
    (20, "River reaching the sea", "Long journey completed", "400x400", "SVG + PNG @2x", "30-day streak / major milestone"),
    (21, "Open gate to garden", "First devotional completed", "400x400", "SVG + PNG @2x", "First completion celebration"),
    (22, "Two flames side by side", "Shared a devotional with someone", "400x400", "SVG + PNG @2x", "Share achievement"),
]

row_num3 = 2
for item in empty_states:
    if item[1] is None:
        ws3.cell(row=row_num3, column=1, value=item[0])
        style_category_row(ws3, row_num3, len(headers3))
        row_num3 += 1
    else:
        vals = list(item) + ["To Do"]
        for col_idx, val in enumerate(vals, 1):
            ws3.cell(row=row_num3, column=col_idx, value=val)
            style_body_cell(ws3, row_num3, col_idx)
        ws3.cell(row=row_num3, column=1).alignment = center_alignment
        row_num3 += 1


# ══════════════════════════════════════════════════════════════
# SHEET 4: UI ICON SET
# ══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("UI Icons")
ws4.sheet_properties.tabColor = "2E86AB"

headers4 = ["#", "Category", "Icon Name", "Description / Purpose", "Sizes", "Format", "States", "Status"]
col_widths4 = [5, 20, 24, 40, 18, 10, 25, 12]
for i, w in enumerate(col_widths4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

ws4.append(headers4)
style_header_row(ws4, 1, len(headers4))
ws4.freeze_panes = "A2"

icons = [
    ("NAVIGATION & CHROME", None),
    (1, "Navigation", "Home", "Main tab bar / bottom nav", "16/24/32", "SVG", "Default, Active"),
    (2, "Navigation", "Search", "Search bar, explore section", "16/24/32", "SVG", "Default, Active"),
    (3, "Navigation", "Library / Bookshelf", "Reading list, saved devotionals", "16/24/32", "SVG", "Default, Active"),
    (4, "Navigation", "Profile / User Circle", "Account, settings access", "16/24/32", "SVG", "Default, Active"),
    (5, "Navigation", "Settings / Gear", "Preferences page", "16/24/32", "SVG", "Default"),
    (6, "Navigation", "Back Arrow", "Navigation back", "16/24/32", "SVG", "Default"),
    (7, "Navigation", "Forward Arrow", "Navigation forward", "16/24/32", "SVG", "Default"),
    (8, "Navigation", "Close / X", "Dismiss modals, sheets", "16/24/32", "SVG", "Default"),
    (9, "Navigation", "Menu / Hamburger", "Drawer navigation toggle", "16/24/32", "SVG", "Default, Open"),
    (10, "Navigation", "More / Ellipsis", "Overflow actions menu", "16/24/32", "SVG", "Default"),

    ("READING & CONTENT", None),
    (11, "Reading", "Book (Open)", "Devotional reading mode", "16/24/32", "SVG", "Default"),
    (12, "Reading", "Book (Closed)", "Series/library browsing", "16/24/32", "SVG", "Default"),
    (13, "Reading", "Bookmark (Outline)", "Save for later action", "16/24/32", "SVG", "Default (unsaved)"),
    (14, "Reading", "Bookmark (Filled)", "Already saved state", "16/24/32", "SVG", "Active (saved)"),
    (15, "Reading", "Scroll", "Scripture, Torah reference", "16/24/32", "SVG", "Default"),
    (16, "Reading", "Text Size (Aa)", "Font size toggle", "24/32", "SVG", "Default"),
    (17, "Reading", "Reading Time / Clock", "Estimated reading time", "16/24", "SVG", "Default"),
    (18, "Reading", "Depth 1 (Single Line)", "1-minute immersion indicator", "16/24", "SVG", "Default, Active"),
    (19, "Reading", "Depth 2 (Double Line)", "5-minute immersion indicator", "16/24", "SVG", "Default, Active"),
    (20, "Reading", "Depth 3 (Triple Line)", "15-minute immersion indicator", "16/24", "SVG", "Default, Active"),
    (21, "Reading", "Checkmark / Complete", "Finished devotional", "16/24/32", "SVG", "Default"),
    (22, "Reading", "Progress Circle", "Partial completion indicator", "16/24/32", "SVG", "0-100%"),
    (23, "Reading", "Play", "Audio devotional playback", "16/24/32", "SVG", "Default"),
    (24, "Reading", "Pause", "Audio pause control", "16/24/32", "SVG", "Default"),
    (25, "Reading", "Volume", "Audio volume control", "16/24/32", "SVG", "Default, Muted"),
    (26, "Reading", "Headphones", "Audio mode indicator", "16/24/32", "SVG", "Default"),

    ("DEVOTIONAL MODULES", None),
    (27, "Module", "Cross (Simple)", "Scripture module icon", "16/24/32", "SVG", "Default"),
    (28, "Module", "Speech Bubble / Quote", "Teaching module icon", "16/24/32", "SVG", "Default"),
    (29, "Module", "Lightbulb", "Insight module icon", "16/24/32", "SVG", "Default"),
    (30, "Module", "Pen / Feather Quill", "Story module icon", "16/24/32", "SVG", "Default"),
    (31, "Module", "Bridge", "Bridge module icon", "16/24/32", "SVG", "Default"),
    (32, "Module", "Mirror / Reflection", "Reflection module icon", "16/24/32", "SVG", "Default"),
    (33, "Module", "Praying Hands", "Prayer module icon", "16/24/32", "SVG", "Default"),
    (34, "Module", "Compass", "Takeaway module icon", "16/24/32", "SVG", "Default"),
    (35, "Module", "Aleph (א)", "Vocab / word study module icon", "16/24/32", "SVG", "Default"),
    (36, "Module", "People (Group)", "Comprehension / discussion icon", "16/24/32", "SVG", "Default"),
    (37, "Module", "Book + Magnifier", "Resource module icon", "16/24/32", "SVG", "Default"),
    (38, "Module", "Person Silhouette", "Profile / character study icon", "16/24/32", "SVG", "Default"),

    ("PATHWAYS", None),
    (39, "Pathway", "Moon / Sleep", "Sleep pathway identifier", "16/24/32", "SVG", "Default, Active"),
    (40, "Pathway", "Sun / Sunrise", "Awake pathway identifier", "16/24/32", "SVG", "Default, Active"),
    (41, "Pathway", "Staff / Shepherd's Crook", "Shepherd pathway identifier", "16/24/32", "SVG", "Default, Active"),

    ("ACTIONS & SOCIAL", None),
    (42, "Action", "Share (Export)", "Share devotional externally", "16/24/32", "SVG", "Default"),
    (43, "Action", "Copy Link", "Copy URL to clipboard", "16/24/32", "SVG", "Default, Copied"),
    (44, "Action", "Download", "Save for offline reading", "16/24/32", "SVG", "Default, Downloading"),
    (45, "Action", "Refresh", "Reload content", "16/24/32", "SVG", "Default, Spinning"),
    (46, "Action", "Notification Bell", "Push notification settings", "16/24/32", "SVG", "Default"),
    (47, "Action", "Notification Bell (Dot)", "Unread notification state", "16/24/32", "SVG", "Active (unread)"),
    (48, "Action", "Heart (Outline)", "Like / favorite action", "16/24/32", "SVG", "Default (unliked)"),
    (49, "Action", "Heart (Filled)", "Liked / favorited state", "16/24/32", "SVG", "Active (liked)"),
    (50, "Action", "Send / Paper Plane", "Share to friend directly", "16/24/32", "SVG", "Default"),
    (51, "Action", "Flag", "Report content", "16/24/32", "SVG", "Default"),

    ("USER & ACCOUNT", None),
    (52, "Account", "Log In / Enter", "Sign in action", "16/24/32", "SVG", "Default"),
    (53, "Account", "Log Out / Exit", "Sign out action", "16/24/32", "SVG", "Default"),
    (54, "Account", "Key", "Password, authentication", "16/24/32", "SVG", "Default"),
    (55, "Account", "Shield / Lock", "Privacy, security settings", "16/24/32", "SVG", "Default, Locked, Unlocked"),
    (56, "Account", "Email / Envelope", "Contact, newsletter signup", "16/24/32", "SVG", "Default, Unread"),
    (57, "Account", "Edit / Pencil", "Edit profile, personal notes", "16/24/32", "SVG", "Default"),
    (58, "Account", "Trash", "Delete action (with confirm)", "16/24/32", "SVG", "Default"),
    (59, "Account", "Calendar", "Daily schedule, streak calendar", "16/24/32", "SVG", "Default"),
    (60, "Account", "Flame / Fire", "Streak count indicator", "16/24/32", "SVG", "Default, Active (streak)"),
    (61, "Account", "Trophy / Crown", "Achievement unlocked", "16/24/32", "SVG", "Default"),
    (62, "Account", "Journal / Notebook", "Personal notes feature", "16/24/32", "SVG", "Default"),

    ("SYSTEM & STATUS", None),
    (63, "System", "Wifi", "Online connection status", "16/24", "SVG", "Default"),
    (64, "System", "Wifi Off", "Offline indicator", "16/24", "SVG", "Default"),
    (65, "System", "Warning Triangle", "Error, caution state", "16/24/32", "SVG", "Default"),
    (66, "System", "Info Circle", "Help text, tooltips", "16/24/32", "SVG", "Default"),
    (67, "System", "Check Circle", "Success confirmation", "16/24/32", "SVG", "Default"),
    (68, "System", "X Circle", "Error / failure state", "16/24/32", "SVG", "Default"),
    (69, "System", "Spinner / Loading", "Loading indicator", "16/24/32", "SVG", "Animated"),
    (70, "System", "Dark Mode / Moon", "Theme toggle (to dark)", "24/32", "SVG", "Default"),
    (71, "System", "Light Mode / Sun", "Theme toggle (to light)", "24/32", "SVG", "Default"),
    (72, "System", "Filter / Funnel", "Content filtering", "16/24", "SVG", "Default, Active"),
    (73, "System", "Sort", "Content ordering", "16/24", "SVG", "Default"),
    (74, "System", "Grid View", "Layout toggle (grid)", "16/24", "SVG", "Default, Active"),
    (75, "System", "List View", "Layout toggle (list)", "16/24", "SVG", "Default, Active"),
    (76, "System", "External Link", "Opens in new tab indicator", "16/24", "SVG", "Default"),
    (77, "System", "Expand / Fullscreen", "Immersive reading mode", "24/32", "SVG", "Default"),
    (78, "System", "Collapse / Minimize", "Exit immersive mode", "24/32", "SVG", "Default"),
    (79, "System", "Chevron Down", "Accordion expand, dropdown", "16/24", "SVG", "Default, Rotated"),
    (80, "System", "Chevron Right", "Drill-in, next item", "16/24", "SVG", "Default"),
]

row_num4 = 2
for item in icons:
    if item[1] is None:
        ws4.cell(row=row_num4, column=1, value=item[0])
        style_category_row(ws4, row_num4, len(headers4))
        row_num4 += 1
    else:
        vals = list(item) + ["To Do"]
        for col_idx, val in enumerate(vals, 1):
            ws4.cell(row=row_num4, column=col_idx, value=val)
            style_body_cell(ws4, row_num4, col_idx)
        ws4.cell(row=row_num4, column=1).alignment = center_alignment
        row_num4 += 1


# ══════════════════════════════════════════════════════════════
# SHEET 5: APP ICONS & PWA ASSETS
# ══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("App Icons & PWA")
ws5.sheet_properties.tabColor = "6B8E23"

headers5 = ["#", "Asset Type", "Asset Name", "Dimensions (px)", "Format", "Notes", "Status"]
col_widths5 = [5, 22, 30, 18, 10, 45, 12]
for i, w in enumerate(col_widths5, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

ws5.append(headers5)
style_header_row(ws5, 1, len(headers5))
ws5.freeze_panes = "A2"

pwa_assets = [
    ("FAVICONS", None),
    (1, "Favicon", "favicon.ico", "32 x 32", "ICO", "Multi-size ICO (16+32), browser tab"),
    (2, "Favicon", "favicon.svg", "Scalable", "SVG", "Modern browsers, respects dark/light mode"),
    (3, "Favicon", "favicon-16x16.png", "16 x 16", "PNG", "Legacy fallback"),
    (4, "Favicon", "favicon-32x32.png", "32 x 32", "PNG", "Legacy fallback"),

    ("APP ICONS (PWA / CAPACITOR)", None),
    (5, "App Icon", "apple-touch-icon.png", "180 x 180", "PNG", "iOS home screen, no transparency"),
    (6, "App Icon", "icon-192.png", "192 x 192", "PNG", "Android home screen, manifest"),
    (7, "App Icon", "icon-512.png", "512 x 512", "PNG", "Android splash, PWA install, store listing"),
    (8, "App Icon", "icon-maskable-192.png", "192 x 192", "PNG", "Android adaptive — safe zone inner 80%"),
    (9, "App Icon", "icon-maskable-512.png", "512 x 512", "PNG", "Android adaptive — safe zone inner 80%"),
    (10, "App Icon", "icon-96-shortcut.png", "96 x 96", "PNG", "PWA shortcut actions"),
    (11, "App Icon", "macos-icon-512.png", "512 x 512", "PNG", "macOS dock (Capacitor build)"),
    (12, "App Icon", "windows-tile-150.png", "150 x 150", "PNG", "Windows start tile"),
    (13, "App Icon", "windows-tile-wide-310x150.png", "310 x 150", "PNG", "Windows wide tile"),

    ("PWA SPLASH SCREENS (iOS)", None),
    (14, "Splash", "splash-640x1136.png", "640 x 1136", "PNG", "iPhone SE"),
    (15, "Splash", "splash-750x1334.png", "750 x 1334", "PNG", "iPhone 8"),
    (16, "Splash", "splash-1170x2532.png", "1170 x 2532", "PNG", "iPhone 12/13/14"),
    (17, "Splash", "splash-1290x2796.png", "1290 x 2796", "PNG", "iPhone 14 Pro Max"),
    (18, "Splash", "splash-1320x2868.png", "1320 x 2868", "PNG", "iPhone 15/16 Pro Max"),
    (19, "Splash", "splash-1640x2360.png", "1640 x 2360", "PNG", "iPad 10th gen"),
    (20, "Splash", "splash-1668x2388.png", "1668 x 2388", "PNG", "iPad Pro 11\""),
    (21, "Splash", "splash-2048x2732.png", "2048 x 2732", "PNG", "iPad Pro 12.9\""),

    ("SOCIAL MEDIA PROFILE / BRANDING", None),
    (22, "Brand", "og-default.jpg", "1200 x 630", "JPEG", "Default OG image for pages without custom art"),
    (23, "Brand", "twitter-card-default.jpg", "1200 x 628", "JPEG", "Default Twitter/X card image"),
    (24, "Brand", "logo-dark-bg.svg", "Scalable", "SVG", "Full logo on Tehom Black background"),
    (25, "Brand", "logo-light-bg.svg", "Scalable", "SVG", "Full logo on Scroll White background"),
    (26, "Brand", "logomark-gold.svg", "Scalable", "SVG", "Icon-only mark in God is Gold"),
    (27, "Brand", "wordmark.svg", "Scalable", "SVG", "EUANGELION text wordmark only"),
]

row_num5 = 2
for item in pwa_assets:
    if item[1] is None:
        ws5.cell(row=row_num5, column=1, value=item[0])
        style_category_row(ws5, row_num5, len(headers5))
        row_num5 += 1
    else:
        vals = list(item) + ["To Do"]
        for col_idx, val in enumerate(vals, 1):
            ws5.cell(row=row_num5, column=col_idx, value=val)
            style_body_cell(ws5, row_num5, col_idx)
        ws5.cell(row=row_num5, column=1).alignment = center_alignment
        row_num5 += 1


# ══════════════════════════════════════════════════════════════
# SHEET 6: SUMMARY / ASSET COUNT
# ══════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Summary")
ws6.sheet_properties.tabColor = GOD_IS_GOLD

# Title
ws6.merge_cells("A1:D1")
ws6.cell(row=1, column=1, value="EUANGELION — Master Image Library Summary").font = title_font

ws6.merge_cells("A3:D3")
ws6.cell(row=3, column=1, value="Asset Counts by Category").font = subtitle_font

headers6 = ["Category", "Unique Assets", "With Retina (@2x)", "Sheet"]
col_widths6 = [35, 16, 18, 25]
for i, w in enumerate(col_widths6, 1):
    ws6.column_dimensions[get_column_letter(i)].width = w

row = 5
for i, h in enumerate(headers6, 1):
    ws6.cell(row=row, column=i, value=h)
style_header_row(ws6, row, len(headers6))

summary_data = [
    ("Evergreen Illustrations", 200, 400, "Evergreen Illustrations"),
    ("Background Textures", 6, 6, "Image Size Specs"),
    ("Empty State Illustrations", 12, 24, "Empty States & System"),
    ("Onboarding Illustrations", 5, 10, "Empty States & System"),
    ("Success/Achievement Illustrations", 5, 10, "Empty States & System"),
    ("UI Icons", 80, 80, "UI Icons"),
    ("Favicons", 4, 4, "App Icons & PWA"),
    ("App Icons (PWA/Capacitor)", 9, 9, "App Icons & PWA"),
    ("PWA Splash Screens", 8, 8, "App Icons & PWA"),
    ("Brand Assets (Logo/OG)", 6, 6, "App Icons & PWA"),
    ("Social Templates", 6, 6, "Image Size Specs"),
    ("Placeholder Images", 4, 4, "Image Size Specs"),
]

for i, (cat, count, retina, sheet) in enumerate(summary_data):
    r = row + 1 + i
    ws6.cell(row=r, column=1, value=cat)
    ws6.cell(row=r, column=2, value=count)
    ws6.cell(row=r, column=3, value=retina)
    ws6.cell(row=r, column=4, value=sheet)
    for c in range(1, 5):
        style_body_cell(ws6, r, c)
    ws6.cell(row=r, column=2).alignment = center_alignment
    ws6.cell(row=r, column=3).alignment = center_alignment

# Total row
total_row = row + 1 + len(summary_data)
ws6.cell(row=total_row, column=1, value="TOTAL")
ws6.cell(row=total_row, column=2, value=sum(s[1] for s in summary_data))
ws6.cell(row=total_row, column=3, value=sum(s[2] for s in summary_data))
ws6.cell(row=total_row, column=4, value="")
style_category_row(ws6, total_row, len(headers6))

# Design notes
notes_row = total_row + 3
ws6.merge_cells(f"A{notes_row}:D{notes_row}")
ws6.cell(row=notes_row, column=1, value="Visual Direction Notes").font = subtitle_font

notes = [
    "Style: Sacred Chiaroscuro — light breaking into darkness, Caravaggio-inspired single-source lighting",
    "Palette: Tehom Black (#1A1612) + God is Gold (#C19A6B) + Scroll White (#F7F3ED)",
    "Rare accents: Covenant Burgundy, Gethsemane Olive, Shalom Blue",
    "Treatment: Dithered/halftone hybrid, limited color palettes, contemplative yet bold",
    "Photography: Editorial magazine style — art direction over templates",
    "Icons: 1.5px stroke, round caps/joins, currentColor inheritance, tree-shakeable SVG components",
    "All images optimized: WebP primary, JPEG fallback. Max file sizes per Image Size Specs sheet.",
    "Deliver @1x and @2x for all raster illustration assets. Icons are vector only (SVG).",
]

for i, note in enumerate(notes):
    r = notes_row + 1 + i
    ws6.merge_cells(f"A{r}:D{r}")
    ws6.cell(row=r, column=1, value=f"  {note}").font = body_font


# ══════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════
output_path = "/Users/meltmac/Documents/app-projects/external/euangelion/content/EUANGELION-IMAGE-LIBRARY.xlsx"
wb.save(output_path)
print(f"Saved to: {output_path}")
print(f"Sheets: {wb.sheetnames}")
print(f"Illustrations: {img_num - 1}")
