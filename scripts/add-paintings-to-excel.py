#!/usr/bin/env python3
"""Add reference painting links to the Euangelion Image Library Excel."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# All 200 paintings mapped: (image_num, painting_title, artist, year, url)
paintings = [
    # ── CREATION & NATURE (1-26) ──
    (1, "The Creation of Light", "Gustave Doré", "1866", "https://en.wikipedia.org/wiki/File:Création_de_la_Lumière.jpg"),
    (2, "The Flight into Egypt", "Adam Elsheimer", "1609", "https://en.wikipedia.org/wiki/The_Flight_into_Egypt_(Elsheimer)"),
    (3, "Sunrise with Sea Monsters", "J.M.W. Turner", "c.1845", "https://en.wikipedia.org/wiki/Sunrise_with_Sea_Monsters"),
    (4, "Evening Landscape with an Aqueduct", "Théodore Géricault", "1818", "https://commons.wikimedia.org/wiki/File:Théodore_Géricault_-_Evening_Landscape_with_an_Aqueduct_-_Google_Art_Project.jpg"),
    (5, "The Solitary Tree", "Caspar David Friedrich", "1822", "https://en.wikipedia.org/wiki/The_Lonely_Tree"),
    (6, "Olive Trees with the Alpilles in the Background", "Vincent van Gogh", "1889", "https://en.wikipedia.org/wiki/Olive_Trees_(Van_Gogh_series)"),
    (7, "The Parable of the Good Samaritan (vineyard detail)", "Pieter Bruegel the Elder", "1567", "https://en.wikipedia.org/wiki/The_Seasons_(Bruegel)"),
    (8, "The Harvesters", "Pieter Bruegel the Elder", "1565", "https://en.wikipedia.org/wiki/The_Harvesters_(Bruegel)"),
    (9, "The Parable of the Sower", "Pieter Bruegel the Elder", "1557", "https://commons.wikimedia.org/wiki/File:Pieter_Bruegel_the_Elder_-_The_Parable_of_the_Sower_-_WGA03374.jpg"),
    (10, "The Barren Fig Tree (Bible illustration)", "Gustave Doré", "1866", "https://commons.wikimedia.org/wiki/File:Gustave_Doré_-_The_Holy_Bible_-_Plate_CXXXVIII.jpg"),
    (11, "Ophelia", "John Everett Millais", "1851–52", "https://en.wikipedia.org/wiki/Ophelia_(painting)"),
    (12, "Wanderer above the Sea of Fog", "Caspar David Friedrich", "1818", "https://en.wikipedia.org/wiki/Wanderer_above_the_Sea_of_Fog"),
    (13, "Christ in the Wilderness", "Ivan Kramskoi", "1872", "https://en.wikipedia.org/wiki/Christ_in_the_Wilderness"),
    (14, "Hagar and the Angel in the Desert", "Giovanni Lanfranco", "c.1616", "https://commons.wikimedia.org/wiki/File:Giovanni_Lanfranco_-_Hagar_in_the_Wilderness_-_WGA12454.jpg"),
    (15, "The Great Wave off Kanagawa (cf. Storm on Sea of Galilee)", "Rembrandt", "1633", "https://en.wikipedia.org/wiki/The_Storm_on_the_Sea_of_Galilee"),
    (16, "The Monk by the Sea", "Caspar David Friedrich", "1808–10", "https://en.wikipedia.org/wiki/The_Monk_by_the_Sea"),
    (17, "A River Landscape with Figures and Cattle", "Aelbert Cuyp", "c.1660", "https://commons.wikimedia.org/wiki/File:Aelbert_Cuyp_-_River_Landscape_with_Horseman_and_Peasants.jpg"),
    (18, "The Deluge", "John Martin", "1834", "https://en.wikipedia.org/wiki/The_Deluge_(Martin)"),
    (19, "Rocky Landscape", "Caspar David Friedrich", "c.1823", "https://commons.wikimedia.org/wiki/File:Caspar_David_Friedrich_-_Rocky_landscape_in_the_Elbe_Sandstone_Mountains_-_Google_Art_Project.jpg"),
    (20, "Moses Before the Burning Bush", "Domenico Fetti", "c.1614", "https://en.wikipedia.org/wiki/Moses_and_the_Burning_Bush#/media/File:Feti,_Domenico_-_Moses_before_the_Burning_Bush.jpg"),
    (21, "Israelites Guided by the Pillar of Fire", "William West", "1845", "https://commons.wikimedia.org/wiki/File:Israelites_guided_by_the_Pillar_of_Fire_-_William_West.jpg"),
    (22, "Landscape with Noah's Thankoffering", "Joseph Anton Koch", "c.1803", "https://commons.wikimedia.org/wiki/File:Joseph_Anton_Koch_006.jpg"),
    (23, "Descent of the Holy Spirit (dove)", "Titian", "c.1545", "https://en.wikipedia.org/wiki/File:Tizian_041.jpg"),
    (24, "Agnus Dei", "Francisco de Zurbarán", "1635–40", "https://en.wikipedia.org/wiki/Agnus_Dei_(Zurbarán)"),
    (25, "Daniel in the Lions' Den (lion study)", "Peter Paul Rubens", "c.1615", "https://en.wikipedia.org/wiki/Daniel_in_the_Lions%27_Den_(Rubens)"),
    (26, "Ganymede (eagle)", "Rubens", "1611–12", "https://commons.wikimedia.org/wiki/File:Peter_Paul_Rubens_-_The_Abduction_of_Ganymede_-_WGA20282.jpg"),

    # ── LIGHT & DARKNESS (27-38) — from completed agent ──
    (27, "Young Girl with a Candle", "Godfried Schalcken", "c.1670", "https://commons.wikimedia.org/wiki/File:Godfried_Schalcken_-_Young_Girl_with_a_Candle_-_WGA20944.jpg"),
    (28, "The Magdalen with the Smoking Flame", "Georges de La Tour", "c.1640", "https://commons.wikimedia.org/wiki/File:Georges_de_La_Tour_-_The_Magdalen_with_the_Smoking_Flame_-_Google_Art_Project.jpg"),
    (29, "Philosopher in Meditation", "Rembrandt van Rijn", "1632", "https://commons.wikimedia.org/wiki/File:Rembrandt_-_The_Philosopher_in_Meditation.jpg"),
    (30, "Buttermere Lake, Cumberland, a Shower", "J.M.W. Turner", "1798", "https://en.wikipedia.org/wiki/Buttermere_Lake,_with_Part_of_Cromackwater,_Cumberland,_a_Shower"),
    (31, "David with the Head of Goliath", "Caravaggio", "c.1610", "https://commons.wikimedia.org/wiki/File:David_with_the_Head_of_Goliath-Caravaggio_(1610).jpg"),
    (32, "Christ before the High Priest", "Gerard van Honthorst", "c.1617", "https://commons.wikimedia.org/wiki/File:Gerard_van_Honthorst_-_Christ_before_the_High_Priest_-_WGA11650.jpg"),
    (33, "Morning in the Riesengebirge", "Caspar David Friedrich", "1810–11", "https://en.wikipedia.org/wiki/Morning_on_the_Riesengebirge"),
    (34, "Moonlit Landscape with River", "Aert van der Neer", "1647", "https://commons.wikimedia.org/wiki/File:Moonlit_Landscape_with_a_View_of_the_New_Amstel_River_and_Castle_Kostverloren_by_Aert_van_der_Neer.JPG"),
    (35, "Bell Rock Lighthouse", "J.M.W. Turner", "1819", "https://commons.wikimedia.org/wiki/File:Joseph_Mallord_William_Turner_-_Bell_Rock_Lighthouse_-_Google_Art_Project.jpg"),
    (36, "Campfire", "Albert Bierstadt", "1863", "https://commons.wikimedia.org/wiki/File:Campfire_(Albert_Bierstadt),_1863.jpg"),
    (37, "The Flight into Egypt (starfield)", "Adam Elsheimer", "1609", "https://en.wikipedia.org/wiki/The_Flight_into_Egypt_(Elsheimer)"),
    (38, "Astronomers Studying an Eclipse", "Antoine Caron", "1571", "https://commons.wikimedia.org/wiki/File:Antoine_Caron_Astronomers_Studying_an_Eclipse.jpg"),

    # ── SACRED OBJECTS & SYMBOLS (39-75) ──
    (39, "Isaiah's Scroll (Dead Sea Scrolls photo)", "Historical artifact", "c.150 BCE", "https://en.wikipedia.org/wiki/Isaiah_scroll"),
    (40, "The Seven Seals (Revelation)", "Albrecht Dürer", "1498", "https://en.wikipedia.org/wiki/File:Dürer_Revelation_Four_Riders.jpg"),
    (41, "Still Life with Bible", "Vincent van Gogh", "1885", "https://en.wikipedia.org/wiki/Still_Life_with_Bible"),
    (42, "Hebrew Torah Scroll (manuscript)", "Historical artifact", "Medieval", "https://commons.wikimedia.org/wiki/File:Torah_and_jad.jpg"),
    (43, "Codex Sinaiticus", "Historical artifact", "4th century", "https://en.wikipedia.org/wiki/Codex_Sinaiticus"),
    (44, "The Supper at Emmaus (bread detail)", "Caravaggio", "1601", "https://en.wikipedia.org/wiki/Supper_at_Emmaus_(Caravaggio,_London)"),
    (45, "The Last Supper (chalice detail)", "Juan de Juanes", "c.1560", "https://en.wikipedia.org/wiki/File:Juan_de_Juanes_-_The_Last_Supper_-_Google_Art_Project.jpg"),
    (46, "Sacrament of the Last Supper", "Salvador Dalí", "1955", "https://en.wikipedia.org/wiki/The_Sacrament_of_the_Last_Supper"),
    (47, "Crucifixion", "Matthias Grünewald", "1512–16", "https://en.wikipedia.org/wiki/Isenheim_Altarpiece"),
    (48, "Cross in the Mountains", "Caspar David Friedrich", "1808", "https://en.wikipedia.org/wiki/Cross_in_the_Mountains"),
    (49, "Man of Sorrows (crown of thorns)", "Albrecht Dürer", "1493", "https://commons.wikimedia.org/wiki/File:Albrecht_Dürer_-_Man_of_Sorrows_-_WGA07017.jpg"),
    (50, "Christ with Crown of Thorns", "Guido Reni", "c.1636", "https://commons.wikimedia.org/wiki/File:Guido_Reni_-_Head_of_Christ_Crowned_with_Thorns_-_WGA19289.jpg"),
    (51, "Christ on the Cross (detail)", "Diego Velázquez", "1632", "https://en.wikipedia.org/wiki/Christ_Crucified_(Velázquez)"),
    (52, "The Holy Women at the Tomb", "Annibale Carracci", "c.1590", "https://commons.wikimedia.org/wiki/File:Annibale_Carracci_-_The_Three_Marys_at_the_Tomb_-_WGA04425.jpg"),
    (53, "Noli Me Tangere (burial cloth)", "Fra Angelico", "c.1440", "https://en.wikipedia.org/wiki/Noli_me_tangere_(Fra_Angelico)"),
    (54, "The Good Shepherd", "Philippe de Champaigne", "c.1650", "https://commons.wikimedia.org/wiki/File:Philippe_de_Champaigne_-_The_Good_Shepherd_-_WGA04715.jpg"),
    (55, "Still Life (earthen vessel)", "Jean-Baptiste-Siméon Chardin", "c.1730", "https://commons.wikimedia.org/wiki/File:Jean_Siméon_Chardin_-_Still_Life_-_The_Kitchen_Table_-_WGA04757.jpg"),
    (56, "God as Architect / Geometer", "Bible Moralisée illumination", "c.1220", "https://en.wikipedia.org/wiki/God_the_Geometer"),
    (57, "Samuel Anointing David", "Dura Europos Synagogue", "c.244 CE", "https://en.wikipedia.org/wiki/Dura-Europos_synagogue"),
    (58, "The Angel Standing in the Sun (incense)", "J.M.W. Turner", "1846", "https://en.wikipedia.org/wiki/The_Angel_Standing_in_the_Sun"),
    (59, "Menorah from Arch of Titus", "Historical relief", "c.82 CE", "https://en.wikipedia.org/wiki/Menorah_(Temple)"),
    (60, "Shofar (Jewish ceremony painting)", "Marc Chagall", "1912–31", "https://en.wikipedia.org/wiki/Solitude_(Chagall)"),
    (61, "Sacrifice of Isaac (altar)", "Caravaggio", "c.1603", "https://en.wikipedia.org/wiki/Sacrifice_of_Isaac_(Caravaggio)"),
    (62, "Baptism of Christ", "Andrea del Verrocchio & Leonardo", "1472–75", "https://en.wikipedia.org/wiki/Baptism_of_Christ_(Verrocchio_and_Leonardo)"),
    (63, "Christ Washing the Feet of the Disciples", "Tintoretto", "c.1556", "https://en.wikipedia.org/wiki/Christ_Washing_the_Feet_of_the_Disciples_(Tintoretto)"),
    (64, "Job on the Dunghill", "Gonzalo Pérez", "c.1440", "https://commons.wikimedia.org/wiki/File:Job_on_the_Dunghill.jpg"),
    (65, "The Goldsmith", "Petrus Christus", "1449", "https://en.wikipedia.org/wiki/A_Goldsmith_in_his_Shop"),
    (66, "Saint Paul (armor detail)", "Rembrandt van Rijn", "1627", "https://commons.wikimedia.org/wiki/File:Rembrandt_-_Apostle_Paul_-_WGA19116.jpg"),
    (67, "Archangel Michael (with sword)", "Guido Reni", "1635", "https://en.wikipedia.org/wiki/Archangel_Michael_(Guido_Reni)"),
    (68, "The Delivery of the Keys", "Pietro Perugino", "1481–82", "https://en.wikipedia.org/wiki/The_Delivery_of_the_Keys_(Perugino)"),
    (69, "The Liberation of Saint Peter", "Raphael", "1514", "https://en.wikipedia.org/wiki/The_Liberation_of_Saint_Peter"),
    (70, "Christ and the Rich Young Ruler", "Heinrich Hofmann", "1889", "https://en.wikipedia.org/wiki/Christ_and_the_Rich_Young_Ruler"),
    (71, "The Last Judgment (scales detail)", "Rogier van der Weyden", "c.1450", "https://en.wikipedia.org/wiki/The_Last_Judgment_(van_der_Weyden)"),
    (72, "The Builders (cornerstone)", "Gustave Doré", "1866", "https://commons.wikimedia.org/wiki/File:Gustave_Doré_-_The_Holy_Bible_-_Plate_I.jpg"),
    (73, "Girl with a Pearl Earring", "Johannes Vermeer", "c.1665", "https://en.wikipedia.org/wiki/Girl_with_a_Pearl_Earring"),
    (74, "The Parable of the Hidden Treasure", "Rembrandt", "1630", "https://commons.wikimedia.org/wiki/File:Rembrandt_-_The_Parable_of_the_Rich_Fool_-_WGA19262.jpg"),
    (75, "The Breakfast Table (still life)", "Willem Claesz. Heda", "1631", "https://commons.wikimedia.org/wiki/File:Willem_Claesz_Heda_002.jpg"),

    # ── HANDS & HUMAN FIGURES (76-95) ──
    (76, "The Creation of Adam (hand of Adam)", "Michelangelo", "c.1512", "https://en.wikipedia.org/wiki/The_Creation_of_Adam"),
    (77, "Praying Hands", "Albrecht Dürer", "1508", "https://en.wikipedia.org/wiki/Praying_Hands_(Dürer)"),
    (78, "Assumption of the Virgin (raised hands)", "Titian", "1516–18", "https://en.wikipedia.org/wiki/Assumption_of_the_Virgin_(Titian)"),
    (79, "Old Woman Praying (Rembrandt's Mother)", "Rembrandt van Rijn", "1629–30", "https://commons.wikimedia.org/wiki/File:Rembrandt_-_An_Old_Woman_Reading_-_WGA19175.jpg"),
    (80, "St. Francis in Ecstasy", "Caravaggio", "c.1594", "https://en.wikipedia.org/wiki/Saint_Francis_of_Assisi_in_Ecstasy_(Caravaggio)"),
    (81, "The Creation of Adam (both hands)", "Michelangelo", "c.1512", "https://en.wikipedia.org/wiki/The_Creation_of_Adam"),
    (82, "St. Jerome Writing", "Caravaggio", "1605–06", "https://en.wikipedia.org/wiki/Saint_Jerome_Writing_(Caravaggio)"),
    (83, "The Sower", "Jean-François Millet", "1850", "https://en.wikipedia.org/wiki/The_Sower_(Millet)"),
    (84, "Christ Washing the Disciples' Feet", "Ford Madox Brown", "1852–56", "https://en.wikipedia.org/wiki/Jesus_Washing_Peter%27s_Feet_(Ford_Madox_Brown)"),
    (85, "The Pilgrim", "Eastman Johnson", "c.1870", "https://commons.wikimedia.org/wiki/File:Eastman_Johnson_-_The_Girl_I_Left_Behind_Me_-_Google_Art_Project.jpg"),
    (86, "The Agony in the Garden", "Andrea Mantegna", "c.1458", "https://en.wikipedia.org/wiki/Agony_in_the_Garden_(Mantegna)"),
    (87, "The Choice of Hercules", "Annibale Carracci", "1596", "https://en.wikipedia.org/wiki/The_Choice_of_Heracles"),
    (88, "Wanderer above the Sea of Fog", "Caspar David Friedrich", "1818", "https://en.wikipedia.org/wiki/Wanderer_above_the_Sea_of_Fog"),
    (89, "Woman at a Window", "Caspar David Friedrich", "1822", "https://en.wikipedia.org/wiki/Woman_at_a_Window"),
    (90, "The Good Shepherd (mosaic)", "Mausoleum of Galla Placidia", "c.425 CE", "https://en.wikipedia.org/wiki/Mausoleum_of_Galla_Placidia"),
    (91, "The Walk to Emmaus", "Robert Zünd", "1877", "https://commons.wikimedia.org/wiki/File:Robert_Zünd_-_The_Road_to_Emmaus_-_Google_Art_Project.jpg"),
    (92, "The School of Athens", "Raphael", "1509–11", "https://en.wikipedia.org/wiki/The_School_of_Athens"),
    (93, "St. Jerome in His Study", "Antonello da Messina", "1475", "https://en.wikipedia.org/wiki/St._Jerome_in_His_Study_(Antonello_da_Messina)"),
    (94, "The Weeping Woman", "Pablo Picasso", "1937", "https://en.wikipedia.org/wiki/The_Weeping_Woman"),
    (95, "Christ of Saint John of the Cross", "Salvador Dalí", "1951", "https://en.wikipedia.org/wiki/Christ_of_Saint_John_of_the_Cross"),

    # ── ARCHITECTURE & SPACES (96-112) ──
    (96, "The Golden Gate (icon)", "Byzantine mosaic", "c.1310", "https://en.wikipedia.org/wiki/Chora_Church"),
    (97, "The Broad and Narrow Way", "Charlotte Reihlen", "1866", "https://en.wikipedia.org/wiki/The_Broad_and_the_Narrow_Way"),
    (98, "The Broad and Narrow Way (full)", "Charlotte Reihlen", "1866", "https://en.wikipedia.org/wiki/The_Broad_and_the_Narrow_Way"),
    (99, "Interior of the Pantheon, Rome", "Giovanni Paolo Panini", "c.1734", "https://en.wikipedia.org/wiki/Interior_of_the_Pantheon,_Rome_(Panini)"),
    (100, "The Destruction of the Temple of Jerusalem", "Francesco Hayez", "1867", "https://en.wikipedia.org/wiki/Destruction_of_the_Temple_of_Jerusalem_(Hayez)"),
    (101, "Construction of the Temple at Jerusalem", "Jean Fouquet", "c.1470", "https://commons.wikimedia.org/wiki/File:Fouquet_Building_of_a_Cathedral.jpg"),
    (102, "The Last Supper", "Leonardo da Vinci", "1495–98", "https://en.wikipedia.org/wiki/The_Last_Supper_(Leonardo)"),
    (103, "The Garden of Earthly Delights (left panel)", "Hieronymus Bosch", "c.1500", "https://en.wikipedia.org/wiki/The_Garden_of_Earthly_Delights"),
    (104, "View of Jerusalem", "David Roberts", "1839", "https://commons.wikimedia.org/wiki/File:David_Roberts-IsraelAndPalestine-1842-Jerusalem-V1.jpg"),
    (105, "Christ in Majesty (Pantocrator)", "Cefalù Cathedral mosaic", "c.1148", "https://en.wikipedia.org/wiki/Christ_Pantocrator_(Cefalù)"),
    (106, "The Light of the World", "William Holman Hunt", "1853–54", "https://en.wikipedia.org/wiki/The_Light_of_the_World_(painting)"),
    (107, "Before the Door (Behold I Stand at the Door)", "William Holman Hunt", "1900", "https://en.wikipedia.org/wiki/The_Light_of_the_World_(painting)"),
    (108, "Jacob's Ladder", "William Blake", "c.1800", "https://en.wikipedia.org/wiki/File:Blake_jacobsladder.jpg"),
    (109, "Christ and the Woman of Samaria", "Angelika Kauffman", "1796", "https://commons.wikimedia.org/wiki/File:Angelica_Kauffmann_-_Christ_and_the_Samaritan_Woman_at_the_Well_-_WGA12085.jpg"),
    (110, "The Tower of Babel", "Pieter Bruegel the Elder", "1563", "https://en.wikipedia.org/wiki/The_Tower_of_Babel_(Bruegel)"),
    (111, "Autumn (The Spies with the Grapes of the Promised Land)", "Nicolas Poussin", "1660–64", "https://en.wikipedia.org/wiki/The_Four_Seasons_(Poussin)"),
    (112, "St. Jerome in His Study", "Albrecht Dürer", "1514", "https://en.wikipedia.org/wiki/Saint_Jerome_in_His_Study_(Dürer)"),

    # ── TEXTURES & ABSTRACT (113-128) ──
    (113, "Earth (from Quatre Éléments)", "Arcimboldo", "1566", "https://en.wikipedia.org/wiki/File:Arcimboldo_Earth.jpg"),
    (114, "Water Lilies (pond detail)", "Claude Monet", "1906", "https://en.wikipedia.org/wiki/Water_Lilies_(Monet_series)"),
    (115, "Snow Storm: Hannibal Crossing the Alps", "J.M.W. Turner", "1812", "https://en.wikipedia.org/wiki/Snow_Storm:_Hannibal_and_his_Army_Crossing_the_Alps"),
    (116, "Christ Pantocrator (gold leaf icon)", "Hagia Sophia", "c.1261", "https://en.wikipedia.org/wiki/File:Hagia_Sophia_Christ_Pantocrator.jpg"),
    (117, "St. Joseph the Carpenter (wood detail)", "Georges de La Tour", "c.1642", "https://en.wikipedia.org/wiki/Saint_Joseph_the_Carpenter_(de_La_Tour)"),
    (118, "Book of Kells (manuscript page)", "Irish monks", "c.800", "https://en.wikipedia.org/wiki/Book_of_Kells"),
    (119, "Jeremiah Lamenting the Destruction of Jerusalem", "Rembrandt", "1630", "https://en.wikipedia.org/wiki/Jeremiah_Lamenting_the_Destruction_of_Jerusalem"),
    (120, "The Annunciation (textile detail)", "Fra Angelico", "c.1440", "https://en.wikipedia.org/wiki/Annunciation_(Fra_Angelico,_San_Marco)"),
    (121, "Head of Christ (Crown of Thorns detail)", "Correggio", "c.1525", "https://commons.wikimedia.org/wiki/File:Correggio_-_Head_of_Christ_-_Google_Art_Project.jpg"),
    (122, "Seraph with Burning Coal (Isaiah 6)", "Gustave Doré", "1866", "https://commons.wikimedia.org/wiki/File:Gustave_Doré_-_Isaiah's_Vision.jpg"),
    (123, "Memento Mori / Vanitas (ashes)", "Philippe de Champaigne", "c.1671", "https://en.wikipedia.org/wiki/Vanitas_(Philippe_de_Champaigne)"),
    (124, "Baptism of Christ (water detail)", "Piero della Francesca", "c.1448", "https://en.wikipedia.org/wiki/Baptism_of_Christ_(Piero_della_Francesca)"),
    (125, "Hunters in the Snow (winter)", "Pieter Bruegel the Elder", "1565", "https://en.wikipedia.org/wiki/Hunters_in_the_Snow"),
    (126, "Journey of the Magi", "Sassetta", "c.1435", "https://en.wikipedia.org/wiki/Journey_of_the_Magi_(Sassetta)"),
    (127, "The Hand of God (Creation detail)", "Michelangelo", "c.1512", "https://en.wikipedia.org/wiki/The_Creation_of_Adam"),
    (128, "Ravenna Mosaics (Emperor Justinian)", "San Vitale, Ravenna", "c.547 CE", "https://en.wikipedia.org/wiki/Basilica_of_San_Vitale"),

    # ── SEASONS & TIME (129-136) ──
    (129, "Almond Blossom", "Vincent van Gogh", "1890", "https://en.wikipedia.org/wiki/Almond_Blossoms"),
    (130, "The Harvesters (summer)", "Pieter Bruegel the Elder", "1565", "https://en.wikipedia.org/wiki/The_Harvesters_(Bruegel)"),
    (131, "Autumn Landscape with Four Trees", "Vincent van Gogh", "1885", "https://commons.wikimedia.org/wiki/File:Vincent_van_Gogh_-_Lane_with_Poplars_(1885).jpg"),
    (132, "Winter Landscape with Skaters", "Hendrick Avercamp", "c.1608", "https://en.wikipedia.org/wiki/Winter_Landscape_with_Skaters_(Avercamp)"),
    (133, "Vanitas with Hourglass", "Antonio de Pereda", "c.1640", "https://commons.wikimedia.org/wiki/File:Antonio_de_Pereda_-_Allegory_of_Vanity_-_WGA17165.jpg"),
    (134, "The Astronomer", "Johannes Vermeer", "1668", "https://en.wikipedia.org/wiki/The_Astronomer_(Vermeer)"),
    (135, "Haystack at Giverny (morning)", "Claude Monet", "1891", "https://en.wikipedia.org/wiki/Haystacks_(Monet_series)"),
    (136, "Two Men Contemplating the Moon", "Caspar David Friedrich", "c.1825", "https://en.wikipedia.org/wiki/Two_Men_Contemplating_the_Moon"),

    # ── JOURNEY & MOVEMENT (137-147) ──
    (137, "The Pilgrimage to San Isidro", "Francisco Goya", "c.1821", "https://en.wikipedia.org/wiki/A_Pilgrimage_to_San_Isidro"),
    (138, "Avenue of Poplars in Autumn", "Vincent van Gogh", "1884", "https://commons.wikimedia.org/wiki/File:Van_Gogh_-_Avenue_of_Poplars_in_Autumn_-_Google_Art_Project.jpg"),
    (139, "Mountain of the Holy Cross", "Thomas Moran", "1875", "https://en.wikipedia.org/wiki/Mountain_of_the_Holy_Cross_(Thomas_Moran)"),
    (140, "Valley of the Shadow of Death", "George Inness", "c.1867", "https://commons.wikimedia.org/wiki/File:George_Inness_-_The_Valley_of_the_Shadow_of_Death_-_Google_Art_Project.jpg"),
    (141, "Christ on the Sea of Galilee", "Eugène Delacroix", "1854", "https://en.wikipedia.org/wiki/Christ_on_the_Sea_of_Galilee"),
    (142, "Still Life with Anchor (maritime)", "Dutch School", "c.1700", "https://commons.wikimedia.org/wiki/File:Anchor_Wikimedia_Commons.jpg"),
    (143, "The Gate of Calais", "William Hogarth", "1748", "https://en.wikipedia.org/wiki/O_the_Roast_Beef_of_Old_England"),
    (144, "Landscape with Footbridge", "Albrecht Altdorfer", "c.1518–20", "https://en.wikipedia.org/wiki/Landscape_with_Footbridge"),
    (145, "The Geographer", "Johannes Vermeer", "c.1669", "https://en.wikipedia.org/wiki/The_Geographer"),
    (146, "The Shipwreck", "J.M.W. Turner", "1805", "https://en.wikipedia.org/wiki/The_Shipwreck_(Turner)"),
    (147, "Slave Ship", "J.M.W. Turner", "1840", "https://en.wikipedia.org/wiki/The_Slave_Ship"),

    # ── MAJOR BIBLICAL SCENES (148-184) ──
    (148, "The Creation of Light", "Gustave Doré", "1866", "https://en.wikipedia.org/wiki/File:Création_de_la_Lumière.jpg"),
    (149, "The Garden of Eden", "Thomas Cole", "1828", "https://en.wikipedia.org/wiki/The_Garden_of_Eden_(Thomas_Cole)"),
    (150, "The Fall of Man", "Peter Paul Rubens", "c.1629", "https://en.wikipedia.org/wiki/The_Fall_of_Man_(Rubens)"),
    (151, "The Deluge", "John Martin", "1834", "https://en.wikipedia.org/wiki/The_Deluge_(Martin)"),
    (152, "Abraham and the Stars (God's Promise)", "Gustave Doré", "1866", "https://commons.wikimedia.org/wiki/File:Gustave_Doré_-_Abraham_Journeying_into_the_Land_of_Canaan.jpg"),
    (153, "Jacob Wrestling with the Angel", "Eugène Delacroix", "1861", "https://en.wikipedia.org/wiki/Jacob_Wrestling_with_the_Angel_(Delacroix)"),
    (154, "The Crossing of the Red Sea", "Nicolas Poussin", "1634", "https://en.wikipedia.org/wiki/The_Crossing_of_the_Red_Sea_(Poussin)"),
    (155, "Moses Receiving the Tablets of the Law", "Marc Chagall", "1960–66", "https://en.wikipedia.org/wiki/Chagall_Biblical_Message"),
    (156, "The Pillar of Fire", "Gustave Doré", "1866", "https://commons.wikimedia.org/wiki/File:095.The_Plague_of_Darkness.jpg"),
    (157, "David and Goliath", "Caravaggio", "c.1599", "https://en.wikipedia.org/wiki/David_and_Goliath_(Caravaggio)"),
    (158, "King David Playing the Harp", "Gerard van Honthorst", "1622", "https://en.wikipedia.org/wiki/King_David_Playing_the_Harp_(Honthorst)"),
    (159, "Elijah's Sacrifice", "Gustave Doré", "1866", "https://commons.wikimedia.org/wiki/File:114.Elijah's_Sacrifice_at_Mount_Carmel.jpg"),
    (160, "Elijah and the Still Small Voice", "Gustave Doré", "1866", "https://commons.wikimedia.org/wiki/File:115.Elijah_Nourished_by_an_Angel.jpg"),
    (161, "Daniel in the Lions' Den", "Peter Paul Rubens", "c.1615", "https://en.wikipedia.org/wiki/Daniel_in_the_Lions%27_Den_(Rubens)"),
    (162, "Jonah and the Whale", "Pieter Lastman", "1621", "https://commons.wikimedia.org/wiki/File:Pieter_Lastman_-_Jonah_and_the_Whale_-_Google_Art_Project.jpg"),
    (163, "Isaiah's Vision", "Gustave Doré", "1866", "https://commons.wikimedia.org/wiki/File:Gustave_Doré_-_Isaiah's_Vision.jpg"),
    (164, "The Vision of the Valley of Dry Bones", "Gustave Doré", "1866", "https://en.wikipedia.org/wiki/Valley_of_Dry_Bones"),
    (165, "Adoration of the Shepherds", "Gerrit van Honthorst", "1622", "https://en.wikipedia.org/wiki/Adoration_of_the_Shepherds_(Honthorst)"),
    (166, "Adoration of the Magi", "Gentile da Fabriano", "1423", "https://en.wikipedia.org/wiki/Adoration_of_the_Magi_(Gentile_da_Fabriano)"),
    (167, "The Baptism of Christ", "Andrea del Verrocchio & Leonardo", "1472–75", "https://en.wikipedia.org/wiki/Baptism_of_Christ_(Verrocchio_and_Leonardo)"),
    (168, "Christ in the Wilderness", "Ivan Kramskoi", "1872", "https://en.wikipedia.org/wiki/Christ_in_the_Wilderness"),
    (169, "The Sermon on the Mount", "Carl Bloch", "1877", "https://en.wikipedia.org/wiki/Carl_Bloch"),
    (170, "Christ in the Storm on the Sea of Galilee", "Rembrandt", "1633", "https://en.wikipedia.org/wiki/The_Storm_on_the_Sea_of_Galilee"),
    (171, "Christ Walking on the Water", "Ivan Aivazovsky", "1888", "https://en.wikipedia.org/wiki/Walking_on_water"),
    (172, "The Miracle of the Loaves and Fishes", "Giovanni Lanfranco", "1620", "https://commons.wikimedia.org/wiki/File:Giovanni_Lanfranco_-_Miracle_of_the_Bread_and_Fish_-_WGA12455.jpg"),
    (173, "Christ Blessing the Little Children", "Lucas Cranach the Elder", "1538", "https://en.wikipedia.org/wiki/Christ_Blessing_the_Children"),
    (174, "The Transfiguration", "Raphael", "1516–20", "https://en.wikipedia.org/wiki/Transfiguration_(Raphael)"),
    (175, "Flevit super illam (He wept over it)", "Enrique Simonet", "1892", "https://en.wikipedia.org/wiki/Flevit_super_illam"),
    (176, "The Last Supper", "Leonardo da Vinci", "1495–98", "https://en.wikipedia.org/wiki/The_Last_Supper_(Leonardo)"),
    (177, "Christ in the Garden of Gethsemane", "Heinrich Hofmann", "1890", "https://en.wikipedia.org/wiki/Christ_in_Gethsemane_(Hofmann)"),
    (178, "Christ Crucified", "Diego Velázquez", "1632", "https://en.wikipedia.org/wiki/Christ_Crucified_(Velázquez)"),
    (179, "Crucifixion (hands detail)", "Matthias Grünewald", "1512–16", "https://en.wikipedia.org/wiki/Isenheim_Altarpiece"),
    (180, "The Resurrection", "Carl Bloch", "1881", "https://en.wikipedia.org/wiki/Carl_Bloch"),
    (181, "The Road to Emmaus", "Robert Zünd", "1877", "https://commons.wikimedia.org/wiki/File:Robert_Zünd_-_The_Road_to_Emmaus_-_Google_Art_Project.jpg"),
    (182, "Pentecost", "El Greco", "c.1600", "https://en.wikipedia.org/wiki/Pentecost_(El_Greco)"),
    (183, "The Conversion of Saint Paul", "Caravaggio", "1601", "https://en.wikipedia.org/wiki/Conversion_on_the_Way_to_Damascus_(Caravaggio)"),
    (184, "The New Jerusalem (Apocalypse)", "Gustave Doré", "1866", "https://en.wikipedia.org/wiki/File:Gustave_Doré_-_The_New_Jerusalem.jpg"),

    # ── EMOTIONAL / EXPERIENTIAL (185-192) ──
    (185, "The Weeping Madonna", "Dieric Bouts", "c.1460", "https://en.wikipedia.org/wiki/Mater_Dolorosa_(Bouts)"),
    (186, "The Return of the Prodigal Son", "Rembrandt", "c.1668", "https://en.wikipedia.org/wiki/The_Return_of_the_Prodigal_Son_(Rembrandt)"),
    (187, "The Rest on the Flight into Egypt", "Caravaggio", "c.1597", "https://en.wikipedia.org/wiki/Rest_on_the_Flight_into_Egypt_(Caravaggio)"),
    (188, "The Laughing Fool", "Jacob Cornelisz. van Oostsanen", "c.1500", "https://commons.wikimedia.org/wiki/File:Laughing_Fool.jpg"),
    (189, "St. Teresa in Ecstasy", "Gian Lorenzo Bernini", "1647–52", "https://en.wikipedia.org/wiki/Ecstasy_of_Saint_Teresa"),
    (190, "The Immaculate Conception", "Bartolomé Esteban Murillo", "c.1678", "https://en.wikipedia.org/wiki/The_Immaculate_Conception_of_Los_Venerables"),
    (191, "The Incredulity of Saint Thomas (wounds)", "Caravaggio", "c.1602", "https://en.wikipedia.org/wiki/The_Incredulity_of_Saint_Thomas_(Caravaggio)"),
    (192, "The Penitent Magdalene", "Georges de La Tour", "c.1640", "https://en.wikipedia.org/wiki/Magdalene_with_the_Smoking_Flame"),

    # ── COMMUNITY & CHURCH (193-200) ──
    (193, "The Supper at Emmaus", "Caravaggio", "1601", "https://en.wikipedia.org/wiki/Supper_at_Emmaus_(Caravaggio,_London)"),
    (194, "Pentecost (gathering of disciples)", "El Greco", "c.1600", "https://en.wikipedia.org/wiki/Pentecost_(El_Greco)"),
    (195, "Baptism of Christ", "Piero della Francesca", "c.1448", "https://en.wikipedia.org/wiki/Baptism_of_Christ_(Piero_della_Francesca)"),
    (196, "Cantoria (singers)", "Luca della Robbia", "1431–38", "https://en.wikipedia.org/wiki/Cantoria_(Luca_della_Robbia)"),
    (197, "Christ and the Disciples at Emmaus", "Rembrandt", "1648", "https://en.wikipedia.org/wiki/Supper_at_Emmaus_(Rembrandt,_Louvre)"),
    (198, "The Wedding at Cana", "Paolo Veronese", "1563", "https://en.wikipedia.org/wiki/The_Wedding_at_Cana"),
    (199, "Supper at Emmaus (candlelit)", "Rembrandt", "1629", "https://en.wikipedia.org/wiki/Supper_at_Emmaus_(Rembrandt)"),
    (200, "Confirmation (Seven Sacraments)", "Nicolas Poussin", "1640s", "https://en.wikipedia.org/wiki/Seven_Sacraments_(Poussin)"),
]


# ── Load existing Excel and add painting columns ──
wb_path = "/Users/meltmac/Documents/app-projects/external/euangelion/content/EUANGELION-IMAGE-LIBRARY.xlsx"
wb = openpyxl.load_workbook(wb_path)
ws = wb["Evergreen Illustrations"]

# Add new headers
TEHOM_BLACK = "1A1612"
header_font = Font(name="Inter", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color=TEHOM_BLACK, end_color=TEHOM_BLACK, fill_type="solid")
body_font = Font(name="Inter", size=10, color="333333")
link_font = Font(name="Inter", size=10, color="2E86AB", underline="single")
thin_border = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)
wrap_alignment = Alignment(wrap_text=True, vertical="top")

# New column headers at K, L, M, N
new_headers = ["Reference Painting", "Artist", "Year", "Reference URL"]
new_col_widths = [35, 25, 12, 55]

for i, (header, width) in enumerate(zip(new_headers, new_col_widths)):
    col = 11 + i  # K=11, L=12, M=13, N=14
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col)].width = width

# Build lookup: image_num -> painting data
painting_map = {p[0]: p for p in paintings}

# Walk through rows, match by image number in column A
for row in range(2, ws.max_row + 1):
    num_cell = ws.cell(row=row, column=1)
    if num_cell.value and isinstance(num_cell.value, int):
        img_num = num_cell.value
        if img_num in painting_map:
            _, title, artist, year, url = painting_map[img_num]
            for col, val in [(11, title), (12, artist), (13, year), (14, url)]:
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = link_font if col == 14 else body_font
                cell.alignment = wrap_alignment
                cell.border = thin_border
            # Make URL a hyperlink
            url_cell = ws.cell(row=row, column=14)
            url_cell.hyperlink = url
    else:
        # Category row — extend gold fill
        GOD_IS_GOLD = "C19A6B"
        gold_fill = PatternFill(start_color=GOD_IS_GOLD, end_color=GOD_IS_GOLD, fill_type="solid")
        gold_font = Font(name="Inter", bold=True, size=11, color="1A1612")
        for col in range(11, 15):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.fill = gold_fill
            cell.font = gold_font

wb.save(wb_path)
print(f"Updated: {wb_path}")
print(f"Added {len(paintings)} painting references across 4 new columns (K-N)")
